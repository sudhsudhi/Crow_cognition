import kivy,datetime
kivy.require('1.10.0') # replace with your current kivy version !

from kivy.app import App
from kivy.uix.widget import Widget
from kivy.uix.button import Button
from kivy.graphics import Color, Ellipse,Line,Fbo,Rectangle
from kivy.clock import Clock
from kivy.properties import NumericProperty
import serial,time
import math,sys
import winsound
from openpyxl import Workbook
import multiprocessing
global ser
#ser=serial.Serial('COM4',9600)

# adjust 1. area , 2. time t , 3. port of serial 


class Rect(Widget):
    area =  40000
    width_int=250
    wth=NumericProperty(width_int)  #string 'width' and 'height' are pre-defined kivy thinigs 
    hht=NumericProperty(area/width_int)
    def __init__(self, **kwargs):
        super(Rect,self).__init__(**kwargs)

        '''the variables nextt is used to handle touch event in cases when the crow pecks within the shape. It is used so that 
        even when the touch is within the shape when the shape has disappeared, the touch is ignored and the program doesn't hang.
        It is also used as a time sometimes to make sure the time interval is maintained.

        the variables blank, cleartime and num_trials are used to ensure that if the crow doesn't peck within the shape in the 30s,
        the trial ends and blank screen appears for 5s.

        cleartime is initially sey by update function. It is resetted in two casees , if the crow succesfully touches or if 30 s ends.
        s

        blank also helps in ensuring that the touches received during 5s blank screen is ignored.

        prgrm_start_time acts as a refernce also for excel sheet.


        '''
        self.nextt='start'              #update will make this variable "o" when the program starts
        self.prgrm_start_time = 0       #update will make this variable the program_start_time
        self.cleartime=datetime.timedelta(0,0,0,0)            #this will be set to prgrm_start_time+30 s by update, then increase by 35 seconds(30 s trial + 5 s blank)
        self.blank=False            # This variable is True when the 5s blank screen is running, else it is False
        self.num_trials =0          #number of trials, icreases if the crow pecks correctly or if 30s ends
        print 'wth:' + str(self.wth)
        print 'hht:' + str(self.hht)
        self.peck_no=0
    
    def play_sound(self):
        winsound.PlaySound('crow_caw.wav', winsound.SND_FILENAME)   #works only with wav file , keep this line at the end only or else the shape will disappear only after the sound plays
    
    def on_touch_down(self, touch):   #self refers to the widget, touch to the the tough motion event#on_touch_down is the function called by kivy when you touch. It's name is standard(not variable)
          self.peck_no+=1
          print "NEXTT: " + str(self.nextt)  
          print "blank: " + str(self.blank) 
          print " "

          self.sheet.cell(row= self.peck_no+2, column=1).value = self.peck_no
          
          self.sheet.cell(row= self.peck_no+2,column=2).value = self.num_trials + 1
          self.sheet.cell(row= self.peck_no+2,column=3).value = str(datetime.datetime.now()-self.prgrm_start_time)
          
          #self.sheet.cell(row= self.peck_no+2,column=5).value = str(touch)

          if  self.nextt =='o' and not(self.blank):                      
                    
                x=touch.x-self.center_x
                y=touch.y-self.center_y
                if -self.hht/2<y<self.hht/2 and -self.wth/2<x<self.wth/2:       #make sure the size here is same as that in kv file

                    print 'pecked_inside_rectangle'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Inside'
                    self.boo.save(self.file_name +'.xlsx')
                    #--arduino part1  , part 2 in update function
                    #ser.write('O') #opens gate
                    
                    t= 5  #number of seconds after which , shape reappears.
                    current1=datetime.datetime.now()
                    self.nextt=current1+datetime.timedelta(0,t,0,0)
                                            #datetime.timedelta([days[, seconds[, microseconds[, milliseconds[, minutes[, hours[, weeks]]]]]]]
                    self.cleartime = current1+datetime.timedelta(0,t,0,0) + datetime.timedelta(0,30,0,0)  # resetting cleartime after a succcessful peck by crow
                    with self.canvas.after:     #to add a screen to the  canvas, canvas.after appears above canvas
                            Color(1, 1, 1)
                            Rectangle(pos=self.pos, size=self.size)
                    print 'yoyo'
                    self.play_sound()
                else:
                    print 'pecked_outside_rectangle'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Outside'
                    self.boo.save(self.file_name +'.xlsx')
          else :
                self.sheet.cell(row= self.peck_no+2,column=2).value = "Blank"
                self.sheet.cell(row= self.peck_no+2,column=4).value = 'Blank_Screen'
                self.boo.save(self.file_name +'.xlsx')
                    
    def update(self,dt):        #just check why is it necesaary to add dt, it is not working without dt
        
        if self.nextt=='start':
                self.file_name= raw_input("Enter the name of the excel sheet ( i.e name or date/time of experiment)")
                self.nextt='o'
                self.prgrm_start_time = datetime.datetime.now()
                     #playing sound for the first trial
                self.cleartime= self.prgrm_start_time + datetime.timedelta(0,30,0,0)
                
                self.boo= Workbook()
                self.sheet=self.boo.active
                self.sheet.cell(row= 1,column=1).value = 'S.no. of pecks'
                self.sheet.cell(row= 1,column=2).value = 'Trial no.'
                self.sheet.cell(row= 1,column=3).value = 'Time'
                self.sheet.cell(row= 1,column=4).value = 'Outside/Inside'
                #self.sheet.cell(row= 1,column=5).value = 'Co-ordinates'
                self.boo.save(self.file_name +'.xlsx')





        if datetime.datetime.now().time() > self.cleartime.time() and not(self.blank) :# there is no need of not(self.blank), but it ensures this loop is not called more than once unnnecessarily
                print ' '
                print "5seconds_blank_screen starts," +str(self.num_trials+1) +"th trial finished"
                print "prgrm_start_time: "+ str(self.prgrm_start_time)
                print "cleartime: "+ str(self.cleartime)
                print " "
                with self.canvas.after:     #to add a screen to the  canvas, canvas.after appears above canvas
                            Color(1, 1, 1)
                            Rectangle(pos=self.pos, size=self.size)
                self.blank=True

        if self.blank :
            sss= self.cleartime + datetime.timedelta(0,5,0,0)
            if datetime.datetime.now().time() > sss.time():  #5s blank time over 
                print "5s_blank_screen_ends" +str(self.num_trials+2)+"th trial started"
                print ' '
                
                self.canvas.after.clear()
                
                self.cleartime = self.cleartime + datetime.timedelta(0,35,0,0)  #resetting cleartime,clear time increase by 35s
                self.blank=False
                self.num_trials+=1  #keep this here only, as it ensures 5 s gap

        if self.nextt!='o':
            if datetime.datetime.now().time() > self.nextt.time():
                self.canvas.after.clear()
                
                self.nextt='o'
                #--arduino part2
                #ser.write('C') #closes gate
                print str(self.num_trials+1) +"th trial finished"
                print ' '
                print "5s_blank_screen_ends" +str(self.num_trials+2)+"th trial started"
                print ' '
                self.num_trials+=1 #keep this here only, as it door is closed only after that is the program stopped when num_trials=20
        
        if self.num_trials == 20:
            print "20 trials executed, program ending"
            sys.exit()      #program ends

        

from kivy.lang import Builder
kv_string= '''
#:kivy 1.10.0
<Rect>:
    canvas.before:
        Color:
            rgba: 1, 1, 1 , 1
        Rectangle:
            pos: self.pos
            size: self.size
    canvas:        
        Color:
            rgb:1,0,0
        Rectangle:
            pos: self.center_x-self.wth/2 , self.center_y-self.hht/2
            size: self.wth, self.hht
'''

Builder.load_string(kv_string)
    
class PreTrainApp(App):
    def build(self):
        re=Rect()
        Clock.schedule_interval(re.update, 1.0 / 60.0)
        return re
    
if __name__ == '__main__':
    PreTrainApp().run()
