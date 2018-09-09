import kivy,datetime
kivy.require('1.10.0') # replace with your current kivy version !
from kivy.core.window import Window
from kivy.config import Config
from kivy.app import App
from kivy.uix.widget import Widget
from kivy.uix.button import Button
from kivy.graphics import Color, Ellipse,Line,Fbo,Rectangle,Triangle
from kivy.clock import Clock
from kivy.properties import NumericProperty
import serial,time
import math,sys
import winsound
from openpyxl import Workbook
import multiprocessing
import random
global ser

#ser=serial.Serial('COM1',9600)

# adjust 1. area , 2. time t , 3. port of serial 


class Shapy(Widget):
    

    def __init__(self, **kwargs):

        # CHANGING FOR DIFFERENT SHAPES: you only have to change  initial variables and self.canvas(remeber to update self.center_xcordi in self.canvas and also inside_out) in positioner and inside_out and also the boundary in positioner(random.uniform)
        super(Shapy,self).__init__(**kwargs)

        '''the variables nextt is used to handle touch event in cases when the crow pecks within the circle. It is used so that 
        even when the touch is within the circle when the circle has disappeared, the touch is ignored and the program doesn't hang.
        It is also used as a time sometimes to make sure the time interval is maintained.
        
        nextt='start' initailly(by __init__), it becomes 'o' when the program starts( by update function), 
        nextt != 'o' for 5 s ( or 't' seconds) after which the crow pecks within the shape. In this time nextt will become a datetime.datetime object and indicates the time(absolute time at which screen should reappear).


        the variables blank, cleartime and num_trials are used to ensure that if the crow doesn't peck within the circle in the 30s,
        the trial ends and blank screen appears for 5s.
        blank = True for 5 s (or whatever will the time be set) after a 30s trial is over. All the other times it is False. It is a boolean.



        cleartime is initially set by update function. It is resetted in two casees , if the crow succesfully touches or if 30 s ends.
        s

        blank also helps in ensuring that the touches received during 5s blank screen is ignored.

        prgrm_start_time acts as a refernce also for excel sheet.

        note that blank and nextt are independent variables and are infact for different purposes.

        Only a touch received when next=='o' and blank == False will be checked for if it is within or outside the shape, because (if nextt != 'o' it means it is not even 5s after the last correct touch) or (if blank = True it means it is not even 5s after the last 30s trial ended) 
        
        
        '''
        self.nextt='start'              #update will make this variable "o" when the program starts
        self.prgrm_start_time = 0       #update will make this variable the program_start_time
        self.cleartime=datetime.timedelta(0,0,0,0)            #this will be set to prgrm_start_time+30 s by update, then increase by 35 seconds(30 s trial + 5 s blank)
        self.blank=False            # This variable is True when the 5s blank screen is running, else it is False
        self.num_trials =0          #the serial number of trials, increases if the crow pecks correctly or if 30s ends
        self.trial_t= 30 #number of seconds for which a trial should run(i.e no. of seconds for which a shape will be shown in a trial )
        self.fail_t = 5 #number of seconds for which blank screen should be shown after a trial in which the crow didn't peck within the shape(i.e failed trial)
        self.touched_t = 15 #number of seconds for which blank screen should be shown after the crow pecks within the shape
        self.peck_no=0
        self.tri_center_x=0 #to act as link between canvas and touch_down_function,
        self.tri_center_y=0 #....
        self.ell_center_x=0 #...
        self.ell_center_y=0
        self.rct_center_x=0
        self.rct_center_y=0
        self.cir_center_x=0
        self.cir_center_y=0
        self.trial_list=[5,5,5,5] #number of trials left of each shape(i.e of the sound associated with that sound), 0th for circle, 1st for ellipse, 2nd for rectangle, 3rd for triangle
        self.sound_shape=-1 # will be changed by positioner. will be 0 if sound for circle is played. 1 for ellipse. 2 for rectangel.3 for triangle.
    
    def positioner(self):  #makes sure that there is atleast one shape in each of the 4 sectors. places the shape randomly within the sector.also plays a sound and edits self.trial_list and self.sound_shape

        
        self.canvas.after.clear()  #to remove black screen, whenever it is placed

        l = random.randint(0,3) #gives 0/1/2/3
        while self.trial_list[l]== 0 : # to check trials are not yet finished
            l = random.randint(0,3)

        self.sound_shape=l
        self.trial_list[self.sound_shape]-=1

        pp=['CIRCLE','ELLIPSE','RECTANGLE','TRIANGLE']
        
        print self.sound_shape, " self.sound_shape"
        print "The sound currently played is linked to the shape : "+ pp[self.sound_shape]
        print "Self.trial_list"+str(self.trial_list)

        if self.sound_shape==0:
          winsound.PlaySound('circle_crow.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        if self.sound_shape==1:
          winsound.PlaySound('ellipse_dog.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        if self.sound_shape==2:
          winsound.PlaySound('rectangle_hawk.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        if self.sound_shape==3:
          winsound.PlaySound('triangle_lion.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
                     

        #cir/ell/rec/tri  must be 0,1,2,3 and unique. This is for division into four sectors.
        #cir = 0 ==> circle in right most quarter  , cir == 3 ==> circle in left most quarter.
        pkol=[0,1,2,3]
        
        cir=random.choice(pkol)
        pkol.remove(cir)
        
        ell=random.choice(pkol)
        pkol.remove(ell)

        rec=random.choice(pkol)
        pkol.remove(rec)

        tri=pkol[0]

        print "cir,ell,rec,tri; " +str(4-cir)+', '+str(4-ell)+', '+str(4-rec)+', '+str(4-tri)+'.(ORDER OF THE SHAPES AS APPEARING ON THE SCREEN)'


        #CIRCLE
        self.area =  40000                  #it is here just to initialize (there is no need of updating, but it's still here)
        self.radius= (self.area/math.pi)**0.5    
        self.cir_center_x=random.uniform(self.center_x+self.width/2-(cir+1)*(self.width/4)  + self.radius ,self.center_x+self.width/2-(cir)*(self.width/4)  -self.radius) #uniform produces a random float
        self.cir_center_y=random.uniform(self.center_y-self.height/2 + self.radius ,self.center_y+self.height/2 -self.radius)

        with self.canvas:     
                    Color(1,0,0)
            
                    Ellipse(pos=(self.cir_center_x - self.radius, self.cir_center_y - self.radius),size=(2*self.radius,2*self.radius))

                


        #ELLIPSE

        self.a=150     #width/2
        self.b= self.area/(self.a*math.pi)

        self.ell_center_x=random.uniform(self.center_x+self.width/2-(ell+1)*(self.width/4)  + self.a ,self.center_x+self.width/2-(ell)*(self.width/4)   -self.a) #uniform produces a random float
        self.ell_center_y=random.uniform(self.center_y-self.height/2 + self.b ,self.center_y+self.height/2 -self.b)

        with self.canvas:     
                    Color(1,0,0)
            
                    Ellipse(pos=(self.ell_center_x - self.a, self.ell_center_y - self.b),size=(2*self.a,2*self.b)) 


        #RECTANGLE
        
        self.wth= 250
        self.hht=self.area/self.wth

        self.rct_center_x=random.uniform(self.center_x+self.width/2-(rec+1)*(self.width/4) + self.wth ,self.center_x+self.width/2 -(rec)*(self.width/4) -  self.wth) #uniform produces a random float
        self.rct_center_y=random.uniform(self.center_y-self.height/2 + self.hht ,self.center_y+self.height/2 -self.hht)

        with self.canvas:     
                    Color(1,0,0)
            
                    Rectangle(pos=(self.rct_center_x-self.wth/2 , self.rct_center_y-self.hht/2),size=(self.wth, self.hht))

     
        # TRIANGLE
        
        self.circum_radius= ((4*(self.area))/(3*(3**0.5)))**0.5

        self.tri_center_x=random.uniform(self.center_x+self.width/2-(tri+1)*(self.width/4) + self.circum_radius ,self.center_x+self.width/2 -(tri)*(self.width/4) -  self.circum_radius) #uniform produces a random float
        self.tri_center_y=random.uniform(self.center_y-self.height/2 + self.circum_radius ,self.center_y+self.height/2 -self.circum_radius)

        with self.canvas:     
                    Color(1,0,0)
            
                    Triangle(points=(self.tri_center_x, self.tri_center_y + self.circum_radius , self.tri_center_x - 0.866025*self.circum_radius , self.tri_center_y -0.5*self.circum_radius  , self.tri_center_x +0.866025*self.circum_radius , self.tri_center_y -0.5*self.circum_radius,))

                

    def inside_out(self,tch):  
        
        # returns 0 for circle, 1 for ellipse, 2 for triangle, 3 for rectangel, -1 if out of every shape
        #tch is the touch object 
             pp=['CIRCLE','ELLIPSE','RECTANGLE','TRIANGLE'] 
             #CIRCLE
             xc=tch.x
             yc=tch.y

             xc1=(float(xc)-self.cir_center_x)**2
             yc1=(float(yc)-self.cir_center_y)**2
             if xc1 +yc1 <= (self.radius)**2:
                self.sheet.cell(row= self.peck_no+2,column=6).value = str(pp[0])
                return 0
          
     
            #ELLIPSE

             xc=tch.x
             yc=tch.y
             xc1=((float(xc)-self.ell_center_x)**2)/(self.a**2)
             yc1=(float(yc)-self.ell_center_y)**2/(self.b**2)
             if xc1 +yc1 <= 1:  
                self.sheet.cell(row= self.peck_no+2,column=6).value = str(pp[1])
                return 1
             
            #RECTANGLE
             x=tch.x-self.rct_center_x
             y=tch.y-self.rct_center_y
             if -self.hht/2<y<self.hht/2 and -self.wth/2<x<self.wth/2 : 
                self.sheet.cell(row= self.peck_no+2,column=6).value = str(pp[2]) 
                return 2
            

            #TRIANGLE
             x=tch.x
             y=tch.y

             x1,y1,x2,y2,x3,y3 = self.tri_center_x,self.tri_center_y + self.circum_radius , self.tri_center_x - 0.866025*self.circum_radius , self.tri_center_y -0.5*self.circum_radius  , self.tri_center_x +0.866025*self.circum_radius , self.tri_center_y -0.5*self.circum_radius
             #print x1,y1,x2,y2,x3,y3
             #print 'x1,y1,x2,y2,x3,y3'
             a=(x-x1)-((y-y1)*((x2-x1)/(y2-y1)))
             a1=(x3-x1)-((y3-y1)*((x2-x1)/(y2-y1)))
             b=(x-x1)-((y-y1)*((x3-x1)/(y3-y1)))
             b1=(x2-x1)-((y2-y1)*((x3-x1)/(y3-y1)))
             c=y-y3   #(y3=y2, the third lines equation is y=y3)
             c1=y1-y3
             if a*a1>=0 and b*b1>=0 and c*c1>=0:  
                self.sheet.cell(row= self.peck_no+2,column=6).value = str(pp[3])
                return 3
           

             return -1  #touched outside shape
    def on_touch_down(self, touch):   #self refers to the widget, touch to the the tough motion event#on_touch_down is the function called by kivy when you touch. It's name is standard(not variable)
          self.peck_no+=1
          

          self.sheet.cell(row= self.peck_no+2, column=1).value = self.peck_no
          self.sheet.cell(row= self.peck_no+2,column=2).value = self.num_trials + 1
          self.sheet.cell(row= self.peck_no+2,column=3).value = str(datetime.datetime.now()-self.prgrm_start_time)
          pp=['CIRCLE','ELLIPSE','RECTANGLE','TRIANGLE']
          self.sheet.cell(row= self.peck_no+2,column=5).value = str(pp[self.sound_shape]) #shape
        

          if  self.nextt =='o' and not(self.blank):                      
                print "self.inside_out(touch): ", str(self.inside_out(touch)) ," (0,1,2,3 for peck within circle,ellipse,rectangle,triangle respectively. -1 if touched outside all.)"

                if self.inside_out(touch)!=-1:      #i.e touched within one of the shapes

                    print 'pecked_inside_circle'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Inside_one_of_the_shapes'
                    self.boo.save(self.file_name +'.xlsx')
                    
                    
                    
                    if self.inside_out(touch)==self.sound_shape:  #touched within right shape
                       print "Opening door..."
                      #ser.write('O') #opens gate
                    
                                       
                    
                    current1=datetime.datetime.now()
                    self.nextt=current1+datetime.timedelta(0,self.touched_t,0,0)
                                            #datetime.timedelta([days[, seconds[, microseconds[, milliseconds[, minutes[, hours[, weeks]]]]]]]
                    self.cleartime = current1+datetime.timedelta(0,self.touched_t,0,0) + datetime.timedelta(0,self.trial_t,0,0)  # resetting cleartime after a succcessful peck by crow
                    self.canvas.clear()
                    print 'yoyo'
                    
                else:
                    print 'pecked_outside_the_correct_shape'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Outside_all_the_shape'
                    self.boo.save(self.file_name +'.xlsx')
          else :
                self.sheet.cell(row= self.peck_no+2,column=2).value = "Blank_screen"
                self.sheet.cell(row= self.peck_no+2,column=4).value = 'Blank_Screen'
                self.boo.save(self.file_name +'.xlsx')

                    
    def update(self,dt):        #just check why is it necesaary to add dt, it is not working without dt
        self.canvas.before.clear()
        with self.canvas.before:     #self.canvas.before is updated every time to update tab when tab is resized.    
            Color(1,1,1)
            Rectangle(pos=self.pos,size=self.size)
   
        if self.nextt=='start':
               

                self.file_name =  raw_input("Enter the name of the excel sheet [ i.e name or date/time of experiment]")
                self.positioner()
                self.nextt='o'
                self.prgrm_start_time = datetime.datetime.now()
               
                self.cleartime= self.prgrm_start_time + datetime.timedelta(0,self.trial_t,0,0)            
                self.boo= Workbook()
                self.sheet=self.boo.active
                self.sheet.cell(row= 1,column=1).value = 'S.no. of pecks'
                self.sheet.cell(row= 1,column=2).value = 'Trial no.'
                self.sheet.cell(row= 1,column=3).value = 'Time'
                self.sheet.cell(row= 1,column=4).value = 'PeckedOutside/Inside'
                self.sheet.cell(row= 1,column=5).value = 'Shape_whose_corresponding_sound is being_played'
                self.sheet.cell(row= 1,column=6).value = 'Shape on which crow pecked.'
                
                #self.sheet.cell(row= 1,column=5).value = 'Co-ordinates'
                self.boo.save(self.file_name +'.xlsx')





        if datetime.datetime.now().time() > self.cleartime.time() and not(self.blank) :# there is no need of not(self.blank), but it ensures this loop is not called more than once unnnecessarily
                print ' '
                print "5seconds_blank_screen starts," +str(self.num_trials+1) +"th trial finished"
                print "prgrm_start_time: "+ str(self.prgrm_start_time)
                print "cleartime: "+ str(self.cleartime)
                print " "
                self.canvas.clear()

                self.blank=True

        if self.blank :
            sss= self.cleartime + datetime.timedelta(0,self.fail_t,0,0)
            if datetime.datetime.now().time() > sss.time():  #5s blank time over 
                print "5s_blank_screen_ends" +str(self.num_trials+2)+"th trial started"
                print ' '
                
                self.positioner()                              
                self.cleartime = self.cleartime + datetime.timedelta(0,self.fail_t+self.trial_t,0,0)  #resetting cleartime,clear time increase by 35s
                self.blank=False
                self.num_trials+=1  #keep this here only, as it ensures 5 s gap

        if self.nextt!='o':
            if datetime.datetime.now().time() > self.nextt.time():
                
                                
                self.nextt='o'
                #--arduino part2
               

                self.positioner()   # this should be after the door closes, because self.positioner changess self.trial_state
                print str(self.num_trials+1) +"th trial finished"
                print ' '
                print "15s_blank_screen_ends, " +str(self.num_trials+2)+"th trial started"
                print ' '
                self.num_trials+=1 #keep this here only, as it door is closed only after that is the program stopped when num_trials=20
        
        if self.num_trials == 20:
            print "20 trials executed, program ending"
            sys.exit()      #program ends

        


    
class TestApp(App):
    def build(self):
        shape=Shapy()
        Clock.schedule_interval(shape.update, 1.0 / 60.0)
        return shape
    
if __name__ == '__main__':
    Config.set('graphics','borderless', 0)   #not necessary all the time, but sometimes required to prevent borders from disappearing.Also full screens the window so that first trail is not cluttered.
    Config.write()                            #NOTE : CHANGING Config from one python file will affect running another python file even if it the other python file doesn't change config.
                                                #also editing window_state or fullscreen of graphics hangs the code.
    TestApp().run()
