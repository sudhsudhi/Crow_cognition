import kivy,datetime
kivy.require('1.10.0') # replace with your current kivy version !

from kivy.app import App
from kivy.uix.widget import Widget
from kivy.uix.button import Button
from kivy.graphics import Color, Ellipse,Line,Fbo,Rectangle
from kivy.clock import Clock
from kivy.properties import NumericProperty
import serial,time
import math,sys
import winsound
from openpyxl import Workbook
import multiprocessing
import random
global ser
ser=serial.Serial('COM1',9600)

# adjust 1. area , 2. time t , 3. port of serial 


class Circly(Widget):
    area =  40000
    radius= NumericProperty((area/math.pi)**0.5)

    def __init__(self, **kwargs):
        super(Circly,self).__init__(**kwargs)

        '''the variables nextt is used to handle touch event in cases when the crow pecks within the circle. It is used so that 
        even when the touch is within the circle when the circle has disappeared, the touch is ignored and the program doesn't hang.
        It is also used as a time sometimes to make sure the time interval is maintained.
        
        nextt='start' initailly(by __init__), it becomes 'o' when the program starts( by update function), 
        nextt != 'o' for 5 s ( or 't' seconds) after which the crow pecks within the shape. In this time nextt will become a datetime.datetime object and indicates the time(absolute time at which screen should reappear).


        the variables blank, cleartime and num_trials are used to ensure that if the crow doesn't peck within the circle in the 30s,
        the trial ends and blank screen appears for 5s.
        blank = True for 5 s (or whatever will the time be set) after a 30s trial is over. All the other times it is False. It is a boolean.



        cleartime is initially set by update function. It is resetted in two casees , if the crow succesfully touches or if 30 s ends.
        s

        blank also helps in ensuring that the touches received during 5s blank screen is ignored.

        prgrm_start_time acts as a refernce also for excel sheet.

        note that blank and nextt are independent variables and are infact for different purposes.

        Only a touch received when next=='o' and blank == False will be checked for if it is within or outside the shape, because (if nextt != 'o' it means it is not even 5s after the last correct touch) or (if blank = True it means it is not even 5s after the last 30s trial ended) 
        
        
        '''
        self.nextt='start'              #update will make this variable "o" when the program starts
        self.prgrm_start_time = 0       #update will make this variable the program_start_time
        self.cleartime=datetime.timedelta(0,0,0,0)            #this will be set to prgrm_start_time+30 s by update, then increase by 35 seconds(30 s trial + 5 s blank)
        self.blank=False            # This variable is True when the 5s blank screen is running, else it is False
        self.num_trials =0          #the serial number of trials, increases if the crow pecks correctly or if 30s ends
        print 'RADIUS:' + str(self.radius)
        self.trial_t= 30 #number of seconds for which a trial should run(i.e no. of seconds for which a shape will be shown in a trial )
        self.fail_t = 5 #number of seconds for which blank screen should be shown after a trial in which the crow didn't peck within the shape(i.e failed trial)
        self.touched_t = 15 #number of seconds for which blank screen should be shown after the crow pecks within the shape
        self.peck_no=0
        self.center_xcordi=0 #to act as link between canvas and touch_down_function, self.center_x would give the center of the screen(widget) not the circle
        self.center_ycordi=0 # this and the above varibales are updated by positioner function
        self.sound_list=[10,5,5] # for randomly producing 10 correct and 10 incorrect sounds.
    def positioner(self):   #randomly places the shape anywhere and plays the sound randomly based on sound_list
        

        self.center_xcordi=random.uniform(self.center_x-self.width/2 + self.radius ,self.center_x+self.width/2-self.radius) #uniform produces a random float
        self.center_ycordi=random.uniform(self.center_y-self.height/2 + self.radius ,self.center_y+self.height/2 -self.radius)

        with self.canvas:     
            Color(1,0,0)
            
            Ellipse(pos=(self.center_xcordi - self.radius, self.center_ycordi - self.radius),size=(2*self.radius,2*self.radius))

        k = random.randint(0,2) #gives 0/1/2

        while self.sound_list[k]==0:
            k = random.randint(0,2)
        self.sound_list[k]-=1
        self.ki=k
        if k==0:    
            winsound.PlaySound('correct_sound.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        
        elif k==1:    
            winsound.PlaySound('incorrect_sound1.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        elif k==2:    
            winsound.PlaySound('incorrect_sound2.wav', winsound.SND_FILENAME | winsound.SND_ASYNC)   # winsound.SND_ASYNC makes sure sound is asynchronous with the program. That is program doesn't block till sound is fully played.
        
        print str(self.sound_list)
        print "no. of trials left : correct="+str(self.sound_list[0])+", incorrect1="+str(self.sound_list[1])+", incorrect2="+str(self.sound_list[2])

    def on_touch_down(self, touch):   #self refers to the widget, touch to the the tough motion event#on_touch_down is the function called by kivy when you touch. It's name is standard(not variable)
          self.peck_no+=1
          print "NEXTT: " + str(self.nextt)  
          print "blank: " + str(self.blank) 
          print " "

          self.sheet.cell(row= self.peck_no+2, column=1).value = self.peck_no
          self.sheet.cell(row= self.peck_no+2,column=2).value = self.num_trials + 1
          self.sheet.cell(row= self.peck_no+2,column=3).value = str(datetime.datetime.now()-self.prgrm_start_time)

          if self.ki==0:
            tro="Correct"

          if self.ki==1:
            tro="InCorrect1"

          if self.ki==2:
            tro="InCorrect2"  

          
          #self.sheet.cell(row= self.peck_no+2,column=5).value = str(touch)

          if  self.nextt =='o' and not(self.blank):                      
                xc=touch.x
                yc=touch.y
                xc1=(float(xc)-self.center_xcordi)**2
                yc1=(float(yc)-self.center_ycordi)**2


                if xc1 +yc1 <= (self.radius)**2:      #make sure the size here is same as that in kv file

                    print 'pecked_inside_circle'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Inside'
                    self.sheet.cell(row= self.peck_no+2,column=5).value = str(tro)
                    self.boo.save(self.file_name +'.xlsx')
                    #--arduino part1  , part 2 in update function
                    ser.write('O') #opens gate
                    
                    
                    current1=datetime.datetime.now()
                    self.nextt=current1+datetime.timedelta(0,self.touched_t,0,0)
                                            #datetime.timedelta([days[, seconds[, microseconds[, milliseconds[, minutes[, hours[, weeks]]]]]]]
                    self.cleartime = current1+datetime.timedelta(0,self.touched_t,0,0) + datetime.timedelta(0,self.trial_t,0,0)  # resetting cleartime after a succcessful peck by crow
                    self.canvas.clear()
                    print 'yoyo'
                    
                else:
                    print 'pecked_outside'
                    self.sheet.cell(row= self.peck_no+2,column=4).value = 'Outside'
                    self.sheet.cell(row= self.peck_no+2,column=5).value = str(tro)
                    self.boo.save(self.file_name +'.xlsx')
          else :
                self.sheet.cell(row= self.peck_no+2,column=2).value = "Blank_screen"
                self.sheet.cell(row= self.peck_no+2,column=4).value = 'Blank_Screen'
                self.boo.save(self.file_name +'.xlsx')
                    
    def update(self,dt):        #just check why is it necesaary to add dt, it is not working without dt
        
        if self.nextt=='start':

                self.file_name= raw_input("Enter the name of the excel sheet ( i.e name or date/time of experiment)")
                self.positioner()
                self.nextt='o'
                self.prgrm_start_time = datetime.datetime.now()
                     #playing sound for the first trial
                self.cleartime= self.prgrm_start_time + datetime.timedelta(0,self.trial_t,0,0)
                
                self.boo= Workbook()
                self.sheet=self.boo.active
                self.sheet.cell(row= 1,column=1).value = 'S.no. of pecks'
                self.sheet.cell(row= 1,column=2).value = 'Trial no.'
                self.sheet.cell(row= 1,column=3).value = 'Time'
                self.sheet.cell(row= 1,column=4).value = 'PeckedOutside/Inside'
                self.sheet.cell(row= 1,column=5).value = 'Correct/Incorrect_trial'
                #self.sheet.cell(row= 1,column=5).value = 'Co-ordinates'
                self.boo.save(self.file_name +'.xlsx')





        if datetime.datetime.now().time() > self.cleartime.time() and not(self.blank) :# there is no need of not(self.blank), but it ensures this loop is not called more than once unnnecessarily
                print ' '
                print "5seconds_blank_screen starts," +str(self.num_trials+1) +"th trial finished"
                print "prgrm_start_time: "+ str(self.prgrm_start_time)
                print "cleartime: "+ str(self.cleartime)
                print " "
                self.canvas.clear()

                self.blank=True

        if self.blank :
            sss= self.cleartime + datetime.timedelta(0,self.fail_t,0,0)
            if datetime.datetime.now().time() > sss.time():  #5s blank time over 
                print "5s_blank_screen_ends" +str(self.num_trials+2)+"th trial started"
                print ' '
                
                self.positioner()                              
                self.cleartime = self.cleartime + datetime.timedelta(0,self.fail_t+self.trial_t,0,0)  #resetting cleartime,clear time increase by 35s
                self.blank=False
                self.num_trials+=1  #keep this here only, as it ensures 5 s gap

        if self.nextt!='o':
            if datetime.datetime.now().time() > self.nextt.time():
                
                self.positioner()
                
                self.nextt='o'
                #--arduino part2
                ser.write('C') #closes gate
                print str(self.num_trials+1) +"th trial finished"
                print ' '
                print "15s_blank_screen_ends, " +str(self.num_trials+2)+"th trial started"
                print ' '
                self.num_trials+=1 #keep this here only, as it door is closed only after that is the program stopped when num_trials=20
        
        if self.num_trials == 20:
            print "20 trials executed, program ending"
            sys.exit()      #program ends

        

from kivy.lang import Builder
kv_string= '''
#:kivy 1.10.0
<Circly>:
	canvas.before:
        Color:
            rgba: 1, 1, 1 , 1
        Rectangle:
            pos: self.pos
            size: self.size
	
'''
#points: self.center_x, self.center_y + self.circum_radius , self.center_x - 0.866025*self.circum_radius , self.center_y -0.5*self.circum_radius  , self.center_x +0.866025*self.circum_radius , self.center_y -0.5*self.circum_radius

Builder.load_string(kv_string)
    
class PreTrainApp(App):
    def build(self):
        cir=Circly()
        Clock.schedule_interval(cir.update, 1.0 / 60.0)
        return cir
    
if __name__ == '__main__':
    PreTrainApp().run()
